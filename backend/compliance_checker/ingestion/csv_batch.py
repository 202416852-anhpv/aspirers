"""
compliance_checker/ingestion/csv_batch.py — Parser CSV/XLSX cho batch xử lý hàng loạt.

⚠️ File mẫu Google Sheet gốc trong brief mục 4.4 vẫn chưa fetch được (cần đăng nhập) — nhưng
nhóm đã có 1 file THẬT từ BGK: `test_pdf_images_link/design_samples_template.xlsx` (30 test
case, cột `no, design, target_market, platform, expected_niche, expected_sub_niche,
expected_style, expected_motifs, expected_verdict, expected_violation_type,
expected_violation_detail, expected_confidence, notes`). Đây là bộ CÂU HỎI + ĐÁP ÁN MẪU (BGK
sẽ điền link/path ảnh thật vào cột `design` khi chấm) — parser bên dưới đọc được cả 2: file
input thường (chỉ cột input) VÀ file có kèm cột `expected_*` (tự động bật chế độ "self-grading",
xem `_EXPECTED_COLUMN_ALIASES` + orchestrator.process_batch()).

Input-side linh hoạt alias tên cột vì giám khảo test live có thể tự tạo CSV/XLSX theo ý họ.
Row-level fault isolation bắt buộc: 1 dòng lỗi KHÔNG được làm sập cả batch.

⚠️ Cột "design" có thể là TEXT (link/path, đọc qua values_only) HOẶC ẢNH DÁN TRỰC TIẾP vào ô
(vd giám khảo Ctrl+C từ đâu đó rồi Ctrl+V thẳng vào Excel) — 2 trường hợp này Excel lưu HOÀN
TOÀN khác nhau: ảnh dán là 1 "hình nổi" (floating drawing) neo gần vị trí ô, KHÔNG phải giá trị
cell, nên đọc bằng values_only sẽ ra None cho ô đó dù mắt thường thấy ảnh. parse_xlsx_rows() xử
lý CẢ 2: text trước (ưu tiên), nếu ô trống VÀ có ảnh neo đúng hàng đó -> lấy ảnh nhúng làm
image_base64 luôn (xem _extract_embedded_images_by_row) — KHÔNG cần giám khảo tự upload ảnh lên
đâu đó lấy link, dán thẳng vào Excel vẫn chạy được.

pip install openpyxl (đọc .xlsx)
"""

import base64
import csv
import io

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# Alias các tên cột input hay gặp -> field chuẩn nội bộ. "design" (cột thật trong file mẫu BGK)
# là link/path ảnh — KHÔNG biết trước là url hay file_path, quyết định ở _normalize_row() dựa
# vào nội dung giá trị (bắt đầu bằng http -> url, ngược lại -> file_path).
_INPUT_COLUMN_ALIASES = {
    "file_path": "file_path", "path": "file_path", "filepath": "file_path", "file": "file_path",
    "url": "url", "link": "url", "image_url": "url", "imageurl": "url",
    "design": "_design_raw",
    "platform": "platform",
    "target_country": "target_country", "country": "target_country", "market": "target_country",
    "target_market": "target_country",  # đúng tên cột thật trong file mẫu BGK
    "niche_hint": "niche_hint", "niche": "niche_hint",
}

# Cột "expected_*" (đáp án mẫu BGK) — giữ nguyên dưới dạng field nội bộ (tiền tố "_", giống
# _row_index/_input_ref) để KHÔNG lẫn vào request gửi orchestrator.process_one_design(), chỉ
# dùng để so sánh/self-grading ở tầng orchestrator.process_batch().
_EXPECTED_COLUMN_ALIASES = {
    "expected_niche": "_expected_niche",
    "expected_sub_niche": "_expected_sub_niche",
    "expected_style": "_expected_style",
    "expected_motifs": "_expected_motifs",
    "expected_verdict": "_expected_verdict",
    "expected_violation_type": "_expected_violation_type",
    "expected_violation_detail": "_expected_violation_detail",
    "expected_confidence": "_expected_confidence",
    "notes": "_notes",
}


def _normalize_row(raw_row: dict) -> dict:
    """Map các key alias về field chuẩn — key lạ không alias được thì bỏ qua (không lỗi)."""
    normalized = {}
    for k, v in raw_row.items():
        if k is None:
            continue
        key_clean = str(k).strip().lower().replace(" ", "_")
        std_field = _INPUT_COLUMN_ALIASES.get(key_clean) or _EXPECTED_COLUMN_ALIASES.get(key_clean)
        if std_field and v is not None and str(v).strip():
            normalized[std_field] = str(v).strip()

    design_val = normalized.pop("_design_raw", None)
    if design_val and "file_path" not in normalized and "url" not in normalized:
        if design_val.lower().startswith(("http://", "https://")):
            normalized["url"] = design_val
        else:
            normalized["file_path"] = design_val

    return normalized


def _rows_from_dicts(raw_rows) -> list[dict]:
    """Dùng chung cho cả CSV lẫn XLSX — nhận iterable các dict thô (1 dict/dòng), trả về list
    đã chuẩn hoá + gắn _row_index/_input_ref."""
    rows: list[dict] = []
    for idx, raw_row in enumerate(raw_rows):
        normalized = _normalize_row(raw_row)
        normalized["_row_index"] = idx
        normalized["_input_ref"] = normalized.get("file_path") or normalized.get("url") or f"row_{idx}"
        rows.append(normalized)
    return rows


def parse_csv_rows(csv_content: str) -> list[dict]:
    """
    Trả về list dict đã chuẩn hoá field, giữ nguyên row_index để trace lỗi đúng dòng.
    Dùng csv.DictReader qua io.StringIO — content đã là text (FE tự đọc file CSV thành text
    utf-8-sig trước khi gửi lên, hoặc backend tự decode nếu nhận multipart file — xem main.py).
    """
    reader = csv.DictReader(io.StringIO(csv_content))
    return _rows_from_dicts(reader)


def _extract_embedded_images_by_row(ws) -> dict:
    """
    Trả {native_0indexed_row: raw_bytes} cho mọi ảnh DÁN/NHÚNG trực tiếp vào sheet (KHÔNG phải
    giá trị cell) — xem docstring đầu file. Mỗi hàng chỉ lấy ảnh ĐẦU TIÊN tìm được (đủ cho case
    phổ biến nhất: 1 ảnh/hàng đúng cột design) — không raise nếu 1 ảnh lỗi, bỏ qua ảnh đó.
    """
    result: dict = {}
    for img in getattr(ws, "_images", []) or []:
        try:
            row0 = img.anchor._from.row  # 0-indexed, tính trên TOÀN sheet (kể cả header)
            if row0 not in result:
                result[row0] = img._data()
        except Exception:
            continue
    return result


def parse_xlsx_rows(raw_bytes: bytes) -> list[dict]:
    """
    Đọc file .xlsx (vd design_samples_template.xlsx của BGK) — LUÔN đọc sheet đầu tiên, dòng 1
    là header. Không raise nếu thiếu openpyxl/file lỗi — trả [] để orchestrator/route tự báo
    lỗi rõ ràng (đúng nguyên tắc fallback-safe chung của module), KHÔNG làm sập cả batch route.

    CHỦ Ý dùng read_only=False (mặc định) thay vì read_only=True như bản trước — read_only là
    parser dạng stream, KHÔNG populate được ws._images (ảnh nhúng/dán) nên sẽ bỏ sót hoàn toàn
    trường hợp giám khảo Ctrl+V ảnh thẳng vào cell thay vì gõ link/path.
    """
    if not _OPENPYXL_AVAILABLE:
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]

        all_value_rows = list(ws.iter_rows(values_only=True))
        if not all_value_rows:
            return []
        header = [str(h).strip() if h is not None else "" for h in all_value_rows[0]]
        data_value_rows = all_value_rows[1:]

        image_by_native_row = _extract_embedded_images_by_row(ws)

        rows: list[dict] = []
        for idx, value_row in enumerate(data_value_rows):
            native_row = idx + 1  # +1 vì header chiếm native row 0
            has_text = any(c is not None for c in value_row)
            embedded_bytes = image_by_native_row.get(native_row)
            if not has_text and embedded_bytes is None:
                continue  # hàng thật sự trống (không text, không ảnh dán) -> bỏ qua

            normalized = _normalize_row(dict(zip(header, value_row))) if has_text else {}

            if "file_path" not in normalized and "url" not in normalized and embedded_bytes is not None:
                # Không có text trong cột design (hoặc bất kỳ alias nào) NHƯNG có ảnh dán trực
                # tiếp đúng hàng này -> dùng luôn ảnh đó, KHÔNG bắt giám khảo tự upload lấy link.
                normalized["image_base64"] = base64.b64encode(embedded_bytes).decode("ascii")

            normalized["_row_index"] = idx
            normalized["_input_ref"] = (
                normalized.get("file_path")
                or normalized.get("url")
                or (f"row_{idx} (ảnh dán trực tiếp trong Excel)" if normalized.get("image_base64") else f"row_{idx}")
            )
            rows.append(normalized)
        return rows
    except Exception as e:
        print(f"⚠️ [csv_batch] Không đọc được file xlsx: {e}")
        return []


def decode_csv_bytes(raw_bytes: bytes) -> str:
    """encoding utf-8-sig để đọc đúng file Excel xuất ra có BOM — dùng khi nhận file CSV
    dạng multipart upload thay vì text thô qua JSON body."""
    return raw_bytes.decode("utf-8-sig")


# ---------------------------------------------------------------------------
# OUTPUT — map DesignComplianceResult về hàng CSV/Excel xuất báo cáo tổng hợp.
# Tên cột dưới đây CẦN ĐỐI CHIẾU LẠI với file mẫu BGK thật trước khi demo/nộp bài.
# ---------------------------------------------------------------------------

BATCH_OUTPUT_COLUMNS = [
    "row_index", "input_ref", "status", "niche", "style", "verdict",
    "confidence", "violation_categories", "violation_summary", "fix_suggestion_summary", "error",
    # 2 cột cuối CHỈ có giá trị khi input đi kèm cột "expected_verdict" (vd file mẫu BGK) —
    # rỗng cho batch input thường, không phá vỡ CSV/report cũ (xem orchestrator.process_batch).
    "expected_verdict", "verdict_match",
]


def batch_row_to_csv_dict(row_result: dict) -> dict:
    """
    row_result: 1 phần tử dict trong orchestrator.process_batch()["rows"] (plain dict, KHÔNG
    phải pydantic object — orchestrator.py cố tình trả dict thuần để nhẹ, main.py mới là nơi
    convert sang schemas.BatchReport qua response_model). Convert về 1 dict phẳng đúng
    BATCH_OUTPUT_COLUMNS.
    """
    base = {
        "row_index": row_result.get("row_index"),
        "input_ref": row_result.get("input_ref"),
        "status": row_result.get("status"),
        "niche": "", "style": "", "verdict": "", "confidence": "",
        "violation_categories": "", "violation_summary": "", "fix_suggestion_summary": "",
        "error": row_result.get("error") or "",
        "expected_verdict": "", "verdict_match": "",
    }
    r = row_result.get("result")
    if row_result.get("status") == "OK" and r is not None:
        evidence = r.get("evidence", {})
        base.update({
            "niche": r.get("niche", ""),
            "style": r.get("style", ""),
            "verdict": r.get("final_verdict", ""),
            "confidence": r.get("overall_confidence", ""),
            "violation_categories": ", ".join(evidence.keys()),
            "violation_summary": " | ".join(f"{k}: {v.get('detail', '')}" for k, v in evidence.items() if v.get("detail")),
            "fix_suggestion_summary": " | ".join(f.get("suggestion", "") for f in r.get("fix_suggestions", [])),
        })

    grading = row_result.get("grading")
    if grading:
        base["expected_verdict"] = grading.get("expected", {}).get("expected_verdict", "")
        base["verdict_match"] = "YES" if grading.get("verdict_match") else "NO"
    return base


def write_batch_csv(rows: list[dict]) -> str:
    """Xuất CSV text từ list dict (đã qua batch_row_to_csv_dict) — dùng cho export báo cáo."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=BATCH_OUTPUT_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow({col: row.get(col, "") for col in BATCH_OUTPUT_COLUMNS})
    return buf.getvalue()
